#coding=utf-8

import time, ConfigParser, xlrd, os, glob, shutil
from xlutils.copy import copy
from openpyxl import load_workbook

project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
configPath = project_dir + '/Config/Environment.ini'
configPath2 = project_dir + '/Config/Environment2.ini'
dbpath = project_dir + '/Config/database_config.ini'


# region 文件相关操作
# 读取数据文件并返回为一个字典
def read_data_from_config(filepath):
    dicdata = {}
    config = ConfigParser.ConfigParser()
    config.read(filepath)
    for section in config.sections():
        dicdata.update(dict(config.items(section)))
    return dicdata

# 重写config文件中option的value值
def rewrite_data_in_config(filepath, section, option, value):
    config = ConfigParser.ConfigParser()
    config.read(filepath)
    config.set(section, option, value)
    file = open(filepath, 'w')
    config.write(file)
    file.close()

# 改写excel文件中的单元格数据
def rewrite_data_in_csv(filepath, coord, value):
    wb = load_workbook(filepath)
    ws = wb.active
    ws.cell(coord).value = value
    wb.save(filepath)

def rewrite_data_in_xls(filepath, row, col, value):
    rb = xlrd.open_workbook(filepath,formatting_info=True)
    wb = copy(rb)
    w_sheet = wb.get_sheet(0)
    w_sheet.write(row, col, value)
    wb.save(filepath)
    time.sleep(0.5)

# 读取csv文件中每一行的数据并保存为一个列表返回（列表中每一个元素是一个元祖，该元祖包含了每一行数据的键值对）
def read_data_from_csv(csv_file_path):
    wb = xlrd.open_workbook(csv_file_path)
    sheet = wb.sheet_by_index(0)
    rows = sheet.nrows
    cols = sheet.ncols
    listData = []
    for i in range(1, rows):
        dictTemp = {}
        for j in range(0, cols):
            dictTemp[sheet.cell(0, j).value.encode('utf-8')] = sheet.cell(i, j).value
        listData.append(dictTemp)
    return listData

# 列出文件目录下的文件内容(不递归)
def list_contents(folderPath):
    list = []
    if folder_exists(folderPath):
        list = os.listdir(folderPath)
    return list

# 递归查找某个目录下所有的文件
def list_files(path, allfile=[]):
    filelist = list_contents(path)
    for filename in filelist:
        filepath = os.path.join(path, filename)
        if os.path.isdir(filepath):
            list_files(filepath, allfile)
        else:
            allfile.append(filepath)
    return allfile

# 在指定目录中寻找符合某些名称规律的文件
# 参数strReg表示通配符(包含文件后缀名)，可取值为*表示任意名称，？表示一个字符，[abc]表示匹配字符a,b,c，[!abc]表示匹配a,b,c以外的字符
# 不进行递归查找
# 返回值为文件名称列表
def search_content_in_folder(pattern, folderPath):
    list = []
    if folder_exists(folderPath):
        os.chdir(folderPath)
        list = glob.glob(pattern)
    return list

# 检查文件是否存在
def file_exists(filePath):
    if os.path.exists(filePath) and os.path.isfile(filePath):
        return True
    return False

# 复制文件
def copy_file(source, target):
    if file_exists(source):
        shutil.copy(source, target)

# 移动文件
def move_file(source, target):
    if file_exists(source) and not file_exists(target):
        shutil.move(source, target)

# 重命名文件
def rename_file(source, target):
    if file_exists(source) and not file_exists(target):
        os.rename(source, target)

# 获取文件的绝对路径
def get_file_abs_path(fileName):
    return os.path.abspath(fileName)

# 删除文件
def remove_file(filePath):
    if file_exists(filePath):
        os.remove(filePath)
# endregion

# region 目录相关操作
# 获取指定父目录下最新的子目录地址
def get_latest_folder_path(parentFolderPath):
    list = os.listdir(parentFolderPath)
    dirlist = []
    dictList = {}
    tmpList = []
    for i in range(0, len(list)):
        path = os.path.join(parentFolderPath,list[i])
        if os.path.isdir(path):
            dirlist.append(list[i])

    for i in range(0, len(dirlist)):
        path = os.path.join(parentFolderPath, dirlist[i])
        if os.path.isfile(path):
            continue
        timestamp = os.path.getmtime(path)
        tmpList.append(timestamp)
        dictList[path] = timestamp

    mx = tmpList[0]
    for item in tmpList:
        if mx < item:
            mx = item

    folderPath = ""
    for key in dictList.keys():
        if dictList[key] == mx:
            folderPath = key
    return folderPath

# 检查目录是否存在
def folder_exists(folderPath):
    if os.path.exists(folderPath) and os.path.isdir(folderPath):
        return True
    return False

# 创建文件目录
def create_file_folder(folderPath):
    if not os.path.exists(folderPath):
        os.mkdir(folderPath)

# 删除文件目录
def remove_file_folder(folderPath):
    if folder_exists(folderPath):
        os.rmdir(folderPath)

# 递归查找某个目录下所有的文件夹
def list_folders(path, allfolder=[]):
    filelist = list_contents(path)
    for foldername in filelist:
        folderPath = os.path.join(path, foldername)
        if os.path.isdir(folderPath):
            allfolder.append(folderPath)
            list_folders(folderPath, allfolder)
    return allfolder
# endregion